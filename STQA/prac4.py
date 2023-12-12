import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
heading = ["Studentname", "Subj1", "Subj2", "Subj3", "Total"]
names = ["Maria", "Christina", "Anna", "Alexa", "Rebecca", "Lyla", "Tina", "Mary", "Meredith", "Izzy"]
marks = [70, 45, 78, 98, 67, 83, 72, 91, 77, 55]
row_count = 1

# Write headings
for col in range(1, len(heading) + 1):
    sheet.cell(row=row_count, column=col).value = heading[col - 1]
row_count += 1

# Write data
for i in range(len(marks)):
    sheet.cell(row=row_count, column=1).value = names[i]
    sheet.cell(row=row_count, column=2).value = marks[i]
    sheet.cell(row=row_count, column=3).value = marks[i]
    sheet.cell(row=row_count, column=4).value = marks[i]
    sheet.cell(row=row_count, column=5).value = marks[i] * 3
    row_count += 1

wb.save("students.xlsx")
